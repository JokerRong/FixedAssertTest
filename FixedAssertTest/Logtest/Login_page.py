import time
from selenium import webdriver
from openpyxl import load_workbook
 
driver = webdriver.Chrome()
driver.maximize_window()
url = "http://knowledgemanagement.zosv2.develop.local/#/systemPlannings"
driver.get(url)

def test_login(username,password):    
    try:                                        
        driver.find_element_by_id("Username").clear()
    except:
        pass
    if username != None:
        driver.find_element_by_id("Username").send_keys(username)
    else:
        pass
    driver.find_element_by_id("Password").clear()
    if password != None:  
        driver.find_element_by_id("Password").send_keys(password)
    else:
        pass
    driver.find_element_by_name("button").click()
    time.sleep(3)
    try:
        driver.find_element_by_xpath("/html/body/main/page-top/div/div[2]")
        print("登录成功！！！！")
    except:
        print("登陆失败！！！！")


def get_value(filepath,sheet,row,coulmn):
    workbook = load_workbook(filepath)
    worksheet = workbook[sheet]
    value = worksheet.cell(row ,coulmn ).value
    return value

if __name__ == '__main__':
    for num in range(3,10):
        data = []
        for nums in range(4,6):
            record = get_value("Login_page.xlsx","Sheet9",row =num,coulmn =nums)
            data.append(record)
        test_login(data[0],data[1])