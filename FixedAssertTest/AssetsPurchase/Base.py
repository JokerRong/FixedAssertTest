from selenium import webdriver
import time
import random
from selenium.webdriver.common.keys import Keys
import traceback #获取错误信息 
from openpyxl import load_workbook

driver = webdriver.Chrome()
driver.maximize_window()


def login():
    driver.get("http://fixedassets.zosv2.develop.local")
    time.sleep(1)                                     
    driver.find_element_by_id("Username").clear()
    driver.find_element_by_id("Username").send_keys("SystemAdmin")
    driver.find_element_by_id("Password").clear()
    driver.find_element_by_id("Password").send_keys("Admin")
    driver.find_element_by_name("button").click()
    driver.implicitly_wait(20)
    time.sleep(3)


# 进入资产规划界面
def enter_estate_planing(tabplace):
    # driver.implicitly_wait(5)
    time.sleep(3)
    driver.find_element_by_link_text("资产购置").click()
    time.sleep(1)
    ul = driver.find_element_by_xpath("/html/body/main/div/div/div/section/div/div/div/div/div/ul")
    li_list = ul.find_elements_by_css_selector("li")
    li_list[int(tabplace)].find_element_by_xpath("./button").click()
    time.sleep(1)


# get assetPlanning top table data
def get_top_table_data(filepath):
    wb = load_workbook(filepath)
    sheet_Names = wb.sheetnames
    sheet = wb[sheet_Names[0]]
    record = {}
    
    # get planing_num
    plan_num = time.strftime("%Y%m%d%H%M", time.localtime())
    record['plan_num'] = plan_num
    
    # get ordername
    plan_name = random.randint(1, 200)
    plan_name = '名称' + str(plan_name)
    record['plan_name'] = plan_name

    # get using_department
    using_department = sheet['F2'].value
    record['using_department'] = using_department

    # get plan_type
    plan_type = sheet['B3'].value
    record['plan_type'] = plan_type

    # get plan_date
    plan_date = time.strftime("%Y-%m-%d", time.localtime())
    record['plan_date'] = plan_date

    # get use 
    use = "用途：测试 " + plan_name
    record['use'] = use
    return record
        

# get assetPlanning below table data
def get_below_table_data(filepath):
    wb = load_workbook(filepath)
    sheet_Names = wb.sheetnames
    sheet = wb[sheet_Names[0]]
    record = {}
    records = []
    # get maxrow 行
    total_num = sheet.max_row 
    for row in range(6, total_num+1):
        record['estate_name'] = sheet['A'+ str(row)].value
        record['specification'] = sheet['B' + str(row)].value
        record['estate_type'] = sheet['C' + str(row)].value
        record['number'] = sheet['D'+ str(row)].value
        record['budgetary_unit_price'] = sheet['E'+ str(row)].value
        if sheet['F' + str(row)].value is not None:
            record['supplier'] = sheet['F' + str(row)].value
        else:
            record['supplier'] = ''
        records.append(record)
        record = {}
    return records


# click the button above
def click_top_button(place):
    time.sleep(2)
    ul_top_xpath = "/html/body/main/div/div/div/section/div/div/div/div[2]/div[1]/ul"
    ul_top = driver.find_element_by_xpath(ul_top_xpath)
    li_list_top = ul_top.find_elements_by_css_selector("li")
    driver.implicitly_wait(5)
    li_list_top[int(place)].find_element_by_xpath("./button").click()


# click the botton below
def click_below_button(place):
    time.sleep(2)
    ul_below_xpath = "/html/body/main/div/div/div/section/div/div/div/div[6]/div/ul"
    ul_below = driver.find_element_by_xpath(ul_below_xpath)
    li_list_below = ul_below.find_elements_by_css_selector("li")
    driver.implicitly_wait(5)
    li_list_below[place].find_element_by_xpath("./button").click()


# fill plan form
def fill_plan_form(plan_data, button):
    time.sleep(2)
    form = driver.find_element_by_xpath("/html/body/div[1]/div/div/div")
    form.find_element_by_name("code").send_keys(plan_data['plan_num'])
    form.find_element_by_name("name").send_keys(plan_data['plan_name'])
    # department
    form.find_element_by_name("useDepartmentId").click()
    time.sleep(0.5)
    driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[2]/div[1]/div/div/input[1]").send_keys(plan_data['using_department'])
    driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[2]/div[1]/div/div/input[1]").send_keys(Keys.ENTER)
    # type
    form.find_element_by_name("assetPlanningType").click()
    time.sleep(0.5)
    form.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[2]/div[2]/div/div/input[1]").send_keys(plan_data['plan_type'])
    form.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[2]/div[2]/div/div/input[1]").send_keys(Keys.ENTER)
    # date
    js = 'document.getElementsByName("planningDateTime")[0].removeAttribute("readonly");'
    driver.execute_script(js)
    driver.find_element_by_name("planningDateTime").send_keys(plan_data['plan_date'])
    # purpose
    form.find_element_by_name("mainPurpose").send_keys(plan_data['use'])

    if button is True:
        form.find_element_by_class_name("btn-primary").click()
    else :
        form.find_element_by_class_name("btn-danger").click()
    return plan_data['plan_num']


# input date 
def input_date(name, data):
    js = 'document.getElementsByName("' + name + '")[0].removeAttribute("readonly");' 
    driver.execute_script(js)
    driver.find_element_by_name(name).send_keys(data)


# choose one line
def choose_one_line(code):
    driver.implicitly_wait(10)
    time.sleep(1)
    driver.find_element_by_xpath("//*[contains(text(),'"+ str(code) + "')]" ).click()
    

# 新增资产规划明细表
def fill_detail_table(record, button):
    time.sleep(1)
    driver.find_element_by_name("assetName").send_keys(record['estate_name'])
    driver.find_element_by_name("specificationModel").send_keys(record['specification'])
    driver.find_element_by_name("assetType").click()
    driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[2]/div[1]/div/div/input[1]").send_keys(record['estate_type'])
    time.sleep(0.5)
    driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[2]/div[1]/div/div/input[1]").send_keys(Keys.ENTER)
    driver.find_element_by_name("purchaseQuantity").send_keys(record['number'])
    driver.find_element_by_name("unitPrice").send_keys(record['budgetary_unit_price'])
    
    if button is True:
        driver.find_element_by_class_name("btn-primary").click()
    else :
        driver.find_element_by_class_name("btn-danger").click()


# 资产调入 新增资产申购单
def transfer_planing_bill(code , tender_state, number):
    name = "申购名称" + str(random.randint(1, 200))
    driver.find_element_by_name("name").send_keys(name)
    
    if tender_state is True:
        true_path = "/html/body/div[1]/div/div/div[2]/form/div/div[2]/div/div/div[1]/label/span"
        driver.find_element_by_xpath(true_path).click()
    else:
        false_path = "/html/body/div[1]/div/div/div[2]/form/div/div[2]/div/div/div[2]/label/span"
        driver.find_element_by_xpath(false_path).click()
    
    choose_one_line(code)

    tbody = driver.find_element_by_xpath("/html/body/div[1]/div/div/div[2]/div[2]/div/div[1]/table/tbody")
    tr_list = tbody.find_elements_by_css_selector("tr")
    for num in range (0, int(number)):
        tr_list[num].find_element_by_xpath("./td['"+ str(num) +"']/input").click()
    
    driver.find_element_by_class_name("btn-primary").click()
    time.sleep(1)
    line = driver.find_element_by_xpath("//*[contains(text(),'" + str(code) + "')]")
    order_number = line.find_element_by_xpath("../td[2]").text
    return order_number


# 新增资产购置订单  将申购单新增成为购置单
def add_order(code):
    time.sleep(1)
    table = driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/div[1]/table")
    td = table.find_element_by_xpath("//*[contains(text(),'" + str(code) + "')]")
    td.find_element_by_xpath("../td[2]/input").click()
    driver.find_element_by_name("submitDetectionInterval").click()


# 填充资产购置单的订单信息
def fill_asset_acquisition_detail(upload_file_path):
    time.sleep(1)
    pay_time_path = "/html/body/div[1]/div/div/div/div[2]/div/form/div[3]/div[1]/div/div/div[1]/span/span[1]"
    driver.find_element_by_xpath(pay_time_path).click()
    pay_time_xpath = "/html/body/div[1]/div/div/div/div[2]/div/form/div[3]/div[1]/div/div/input[1]"
    driver.find_element_by_xpath(pay_time_xpath).send_keys(1)
    time.sleep(1)
    driver.find_element_by_xpath(pay_time_xpath).send_keys(Keys.ENTER)

    # fahuoxingshi
    driver.find_element_by_id("shippingForm").send_keys("快递发送")
    input_date("deliveryDate", time.strftime("%Y-%m-%d", time.localtime()))
    input_date("arrivalDate", time.strftime("%Y-%m-%d", time.localtime()))
    
    # uoload file
    driver.find_element_by_id("installationUploadFile").send_keys(upload_file_path)
    
    time.sleep(3)
    driver.find_element_by_class_name("btn-primary").click()
