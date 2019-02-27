import sys,os
from openpyxl import load_workbook


'''
该函数包含两个参数，一个是excel的地址，另一个是excel的sheet名称，返回值：以二维数组的形式返回该sheet中所有的内容
    TableName = "WorkInfor"
    ExcelURL = "StoragesAndSupplies\\Data\\excelquery"
'''
def query_table_all(TableName,ExcelURL):
    wb = load_workbook(os.getcwd() + '\\ZOS.' + ExcelURL +'.xlsx')
    SheetRow = wb[TableName].max_row
    SheetCol = wb[TableName].max_column
    QueryResult = [[0]*SheetCol for row in range(SheetRow)]
    for RowNum in range(1,SheetRow+1):
        for ColNum in range(1,SheetCol+1):
            QueryResult[RowNum-1][ColNum-1] = wb[TableName].cell(row =RowNum,column = ColNum).value
    return QueryResult
'''
该函数包含6个参数，实现了left join的功能
    TableName = "WorkInfor"
    ExcelURL = "StoragesAndSupplies\\Data\\excelquery"
'''
def left_join_table_by(TableNameA,TableNameB,OnAX,OnBX,ExcelURLA,ExcelURLB):
    #获取A表的所有数据
    wbA = load_workbook(os.getcwd() + '\\ZOS.' + ExcelURL +'.xlsx')
    SheetRowA = wb[TableNameA].max_row
    SheetColA = wb[TableNameA].max_column
    ResultA = [[0]*SheetColA for row in range(SheetRowA)]
    for RowNum in range(1,SheetRowA+1):
        for ColNum in range(1,SheetColA+1):
            ResultA[RowNum-1][ColNum-1] = wb[TableNameA].cell(row =RowNum,column = ColNum).value
    #获取B表的所有数据
    wbB = load_workbook(os.getcwd() + '\\ZOS.' + ExcelURLB +'.xlsx')
    SheetRowB = wb[TableNameB].max_row
    SheetColB = wb[TableNameB].max_column
    ResultB = [[0]*SheetColB for row in range(SheetRowB)]
    for RowNum in range(1,SheetRowB+1):
        for ColNum in range(1,SheetColB+1):
            ResultB[RowNum-1][ColNum-1] = wb[TableNameB].cell(row =RowNum,column = ColNum).value
    #找到AX索引
    for RowNum in range(0,SheetRowA):
        for ColNum in range (0,SheetColA):
            if  ResultA[RowNum][ColNum] == OnAX :
                LocationAX = [RowNum,ColNum]
    #找到BX索引
    for RowNum in range(0,SheetRowB):
        for ColNum in range (0,SheetColB):
            if  ResultA[RowNum][ColNum] == OnBX :
                LocationBX = [RowNum,ColNum]  
    #定义结果数组
    Result = [[0]*(SheetColB+SheetColA) for row in range(SheetRowA +SheetRowB)]  
    #根据条件关联合并A\B数组
    for RowNum in range(LocationAX[0],SheetRowA):
        if ResultA[RowNum][LocationAX[0]] == ResultB[RowNum][LocationBX[0]]:
            for ResultColNum in range(0,SheetColA):
                Result[RowNum][ResultColNum] = ResultA[RowNum][ResultColNum] 
            ResultBColNum = 0
            for ResultColNum in range(SheetColA,SheetColB+SheetColA):
                Result[RowNum][ResultColNum] = ResultB[RowNum][ResultBColNum]
                ResultBColNum = ResultBColNum+1
    #整理数组，清除重复列
    for ResultColNum in range (0,len(Result[0])):
        for ResultColNumNext in range(ResultColNum + 1,len(Result[0]) - 1):  
            if  Result[0][ResultColNum] == Result[0][ResultColNumNext]:
                for exchangeNum in range(ResultColNumNext,len(Result[0]) - 1):
                    Result[0][exchangeNum] = Result[0][exchangeNum + 1]
                    Result[0][exchangeNum + 1] = None
                    for exchangeRow in range(1,SheetRowA + SheetRowB):
                        Result[exchangeRow][exchangeNum] = Result[exchangeRow][exchangeNum + 1]
                        Result[exchangeRow][exchangeNum + 1] = None
                for ResultPop in range(0,SheetRowA + SheetRowB):
                    Result[ResultPop].pop(-1)    
    #清除重复行
    ResultPopRow = 1
    while ResultPopRow < len(Result):
        if Result[ResultPopRow][0] == 0:
            Result.pop(ResultPopRow)
            ResultPopRow =ResultPopRow-1
        else:
            ResultPopRow =ResultPopRow+1
    return Result

