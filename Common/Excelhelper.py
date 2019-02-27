import sys,os
from openpyxl import load_workbook

'''
该函数包含两个参数，一个是excel的地址，另一个是excel的sheet名称，返回值：返回该表格最大行
    TableName = "WorkInfor"
    ExcelURL = "StoragesAndSupplies\\Data\\excelquery"
'''
def get_excel_maxRow(TableName,ExcelURL):
    wb = load_workbook(os.getcwd() + '\\ZOS.' + ExcelURL +'.xlsx')
    SheetRow = wb[TableName].max_row
    return SheetRow


'''
该函数包含两个参数，一个是excel的地址，另一个是excel的sheet名称，返回值：返回该表格最大列
    TableName = "WorkInfor"
    ExcelURL = "StoragesAndSupplies\\Data\\excelquery"
'''
def get_excel_maxCol(TableName,ExcelURL):
    wb = load_workbook(os.getcwd() + '\\ZOS.' + ExcelURL +'.xlsx')
    SheetCol = wb[TableName].max_column
    return SheetCol